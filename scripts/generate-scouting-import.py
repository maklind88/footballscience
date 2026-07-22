#!/usr/bin/env python3
"""Generate the lazy-loaded Scouting player database."""

from __future__ import annotations

import json
import hashlib
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_OUTPUT = Path("scouting-import-data.js")
PREVIEW_OUTPUT_NAME = "scouting-import-preview-data.js"
MANIFEST_OUTPUT_NAME = "scouting-import-manifest.js"
ISOLATED_SHEET_NAME = "NWSL (Statsbomb)"
ISOLATED_OUTPUT_NAME = "scouting-statsbomb-data.js"
PREVIEW_RECORD_LIMIT = 50

CORE_HEADERS = {
    "league",
    "season",
    "player",
    "team",
    "team within selected timeframe",
    "position",
    "age",
    "matches",
    "minutes",
    "minutes played",
    "birth country",
    "passport country",
    "height",
    "weight",
}


def clean_text(value, limit=240):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()[:limit]


def slugify(value, fallback="item"):
    slug = re.sub(r"[^a-z0-9]+", "-", clean_text(value).lower()).strip("-")
    return slug or fallback


def to_number(value):
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = clean_text(value)
    if not text or text in {"-", "n/a", "N/A"}:
        return None
    text = text.replace(",", ".").replace("%", "")
    try:
        return float(text)
    except ValueError:
        return None


def normalize_headers(values):
    headers = [clean_text(value, 180) for value in values]
    normalized = [header.lower() for header in headers]
    has_explicit_league = any(header in {"league", "leagie"} for header in normalized)
    has_explicit_season = "season" in normalized

    for index, header in enumerate(headers):
        lowered = header.lower()
        if lowered == "leagie":
            headers[index] = "League"
        elif index == 0 and not has_explicit_league:
            headers[index] = "League"
        elif index == 1 and not has_explicit_season:
            headers[index] = "Season"
    return headers


def metric_group(label):
    text = label.lower()
    if any(token in text for token in ["save", "goal against", "conceded", "gk", "exit"]):
        return "Goalkeeping"
    if any(token in text for token in ["goal", "xg", "shot", "touches in box", "penalty area"]):
        return "Goal threat"
    if any(token in text for token in ["assist", "xa", "key pass", "smart pass", "through pass", "cross"]):
        return "Chance creation"
    if any(token in text for token in ["progressive", "dribble", "carry", "run"]):
        return "Progression"
    if any(token in text for token in ["pass", "received"]):
        return "Passing"
    if any(token in text for token in ["duel", "interception", "recover", "defensive", "aerial"]):
        return "Duels and defending"
    if any(token in text for token in ["yellow", "red", "foul", "loss"]):
        return "Risk"
    return "General"


def metric_direction(label):
    text = label.lower()
    lower_is_better_tokens = [
        "losses",
        "lost",
        "fouls",
        "yellow cards",
        "red cards",
        "goals conceded",
        "goal against",
        "unsuccessful",
        "errors",
    ]
    return "lower" if any(token in text for token in lower_is_better_tokens) else "higher"


def rounded_number(value):
    if value is None:
        return None
    if abs(value) >= 100:
        return round(value, 1)
    return round(value, 4)


def source_sha256(source):
    digest = hashlib.sha256()
    with source.open("rb") as source_file:
        for chunk in iter(lambda: source_file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def serialize_source_value(value):
    if isinstance(value, datetime):
        return value.date().isoformat() if value.time() == datetime.min.time() else value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return clean_text(value, 1000)


def build_isolated_columns(header_row):
    counts = {}
    columns = []
    for index, value in enumerate(header_row):
        label = clean_text(value, 240) or f"Column {index + 1}"
        base = slugify(label, f"column-{index + 1}")
        counts[base] = counts.get(base, 0) + 1
        column_id = base if counts[base] == 1 else f"{base}-{counts[base]}"
        columns.append({"id": column_id, "label": label, "index": index})
    return columns


def build_isolated_database(worksheet, workbook_hash):
    iterator = worksheet.iter_rows(values_only=True)
    header_row = next(iterator, None)
    if not header_row or not any(cell is not None for cell in header_row):
        raise ValueError(f"Isolated source sheet {worksheet.title!r} has no header row.")

    columns = build_isolated_columns(header_row)
    column_index = {column["id"]: column["index"] for column in columns}
    player_index = column_index.get("player")
    player_id_index = column_index.get("player-sbd-id")
    if player_index is None or player_id_index is None:
        raise ValueError(f"Isolated source sheet {worksheet.title!r} is missing Player or Player SBD ID.")

    records = []
    player_ids = set()
    for row in iterator:
        player = row[player_index] if player_index < len(row) else None
        if not clean_text(player, 160):
            continue
        player_id = row[player_id_index] if player_id_index < len(row) else None
        normalized_player_id = clean_text(player_id, 120)
        if not normalized_player_id:
            raise ValueError(f"Isolated source record {player!r} is missing Player SBD ID.")
        if normalized_player_id in player_ids:
            raise ValueError(f"Duplicate Player SBD ID in isolated source: {normalized_player_id}.")
        player_ids.add(normalized_player_id)
        padded_row = list(row[: len(columns)]) + [None] * max(0, len(columns) - len(row))
        records.append([serialize_source_value(value) for value in padded_row])

    return {
        "schema": "football-science-statsbomb-player-database",
        "version": f"football-science-statsbomb-nwsl-v1-{len(records)}-{len(columns)}",
        "source": "NWSL StatsBomb",
        "sourceSheet": worksheet.title,
        "sourceFileSha256": workbook_hash,
        "integrationStatus": "isolated-awaiting-rules",
        "loadedByApplication": False,
        "primaryKeyColumn": "player-sbd-id",
        "columns": columns,
        "records": records,
    }


def write_javascript_payload(output, assignment, payload):
    output.write_text(
        assignment + json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + ";\n",
        encoding="utf-8",
    )


def build_preview_payload(payload, output):
    return {
        "schema": "football-science-scouting-preview",
        "version": f"{payload['version']}-preview-{PREVIEW_RECORD_LIMIT}",
        "source": payload["source"],
        "metricEncoding": payload["metricEncoding"],
        "recordColumns": payload["recordColumns"],
        "metrics": payload["metrics"],
        "sheets": payload["sheets"],
        "importedAt": "",
        "fileName": output.name,
        "totalRecords": len(payload["records"]),
        "records": payload["records"][:PREVIEW_RECORD_LIMIT],
    }


def build_manifest_payload(payload, preview_payload, output, preview_output, isolated_output, isolated_payload):
    return {
        "schema": "football-science-scouting-import-manifest",
        "version": payload["version"],
        "full": {
            "script": output.name,
            "schema": payload["schema"],
            "version": payload["version"],
            "records": len(payload["records"]),
            "metrics": len(payload["metrics"]),
        },
        "preview": {
            "script": preview_output.name,
            "schema": preview_payload["schema"],
            "version": preview_payload["version"],
            "records": len(preview_payload["records"]),
            "metrics": len(preview_payload["metrics"]),
        },
        "isolatedSources": [
            {
                "script": isolated_output.name,
                "schema": isolated_payload["schema"],
                "version": isolated_payload["version"],
                "sourceSheet": isolated_payload["sourceSheet"],
                "records": len(isolated_payload["records"]),
                "columns": len(isolated_payload["columns"]),
                "integrationStatus": isolated_payload["integrationStatus"],
                "loadedByApplication": isolated_payload["loadedByApplication"],
            }
        ],
    }


def main():
    source_arg = sys.argv[1] if len(sys.argv) > 1 else os.environ.get("SCOUTING_PLAYER_DATABASE_SOURCE", "")
    if not source_arg:
        raise SystemExit("Pass scouting player database source as the first argument or set SCOUTING_PLAYER_DATABASE_SOURCE.")
    source = Path(source_arg)
    output = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_OUTPUT
    preview_output = output.with_name(PREVIEW_OUTPUT_NAME)
    manifest_output = output.with_name(MANIFEST_OUTPUT_NAME)
    isolated_output = output.with_name(ISOLATED_OUTPUT_NAME)
    workbook = load_workbook(source, read_only=True, data_only=True)
    workbook_hash = source_sha256(source)

    metric_by_label = {}
    metric_defs = []
    records = []
    sheet_summaries = []
    isolated_payload = None

    def metric_id_for(label):
        if label in metric_by_label:
            return metric_by_label[label]
        base = slugify(label, "metric")
        metric_id = base
        counter = 2
        existing_ids = {metric["id"] for metric in metric_defs}
        while metric_id in existing_ids:
            metric_id = f"{base}-{counter}"
            counter += 1
        metric_by_label[label] = metric_id
        metric_defs.append(
            {
                "id": metric_id,
                "key": base,
                "label": label,
                "group": metric_group(label),
                "direction": metric_direction(label),
            }
        )
        return metric_id

    def value_by(headers, row, names):
        wanted = {name.lower() for name in names}
        for index, header in enumerate(headers):
            if header.lower() in wanted and index < len(row):
                return row[index]
        return None

    for worksheet in workbook.worksheets:
        if worksheet.title == ISOLATED_SHEET_NAME:
            isolated_payload = build_isolated_database(worksheet, workbook_hash)
            continue
        iterator = worksheet.iter_rows(values_only=True)
        header_row = next(iterator, None)
        if not header_row or not any(cell is not None for cell in header_row):
            continue
        headers = normalize_headers(header_row)
        if not any(headers):
            continue

        row_count = 0
        for row in iterator:
            if not row or not any(cell is not None and clean_text(cell) for cell in row):
                continue
            player = clean_text(value_by(headers, row, ["Player"]), 160)
            if not player:
                continue
            league = clean_text(value_by(headers, row, ["League"]), 160) or worksheet.title
            season = clean_text(value_by(headers, row, ["Season"]), 80)
            team = clean_text(value_by(headers, row, ["Team"]), 160)
            team_within_timeframe = clean_text(value_by(headers, row, ["Team within selected timeframe"]), 180)
            position = clean_text(value_by(headers, row, ["Position"]), 80)
            age = rounded_number(to_number(value_by(headers, row, ["Age"])))
            matches = rounded_number(to_number(value_by(headers, row, ["Matches"])))
            minutes = rounded_number(to_number(value_by(headers, row, ["Minutes", "Minutes played"])))
            birth_country = clean_text(value_by(headers, row, ["Birth country"]), 120)
            passport_country = clean_text(value_by(headers, row, ["Passport country"]), 120)
            height = rounded_number(to_number(value_by(headers, row, ["Height"])))
            weight = rounded_number(to_number(value_by(headers, row, ["Weight"])))

            metrics = {}
            for index, header in enumerate(headers):
                if not header or header.lower() in CORE_HEADERS or index >= len(row):
                    continue
                number = rounded_number(to_number(row[index]))
                if number is None:
                    continue
                metrics[metric_id_for(header)] = number

            record_id = "--".join(
                [
                    slugify(player, "player"),
                    slugify(team or team_within_timeframe or league, "team"),
                    slugify(league, "league"),
                    slugify(season, "season"),
                    str(len(records) + 1),
                ]
            )
            records.append(
                [
                    record_id,
                    player,
                    team,
                    team_within_timeframe,
                    league,
                    season,
                    position,
                    age,
                    matches,
                    minutes,
                    birth_country,
                    passport_country,
                    height,
                    weight,
                    metrics,
                ]
            )
            row_count += 1

        sheet_summaries.append({"name": worksheet.title, "rows": row_count})

    if isolated_payload is None:
        raise ValueError(f"Required isolated source sheet {ISOLATED_SHEET_NAME!r} was not found.")

    payload = {
        "schema": "football-science-scouting-import",
        "version": f"scouting-player-database-v1-{len(records)}-{len(metric_defs)}",
        "source": "Scouting player database",
        "metricEncoding": "metric-index-array-v1",
        "recordColumns": [
            "id",
            "player",
            "team",
            "teamWithinTimeframe",
            "league",
            "season",
            "position",
            "age",
            "matches",
            "minutes",
            "birthCountry",
            "passportCountry",
            "height",
            "weight",
            "metrics",
        ],
        "metrics": metric_defs,
        "sheets": sheet_summaries,
        "records": records,
    }

    metric_ids = [metric["id"] for metric in metric_defs]
    for record in records:
        metric_values = record[-1]
        record[-1] = [metric_values.get(metric_id) for metric_id in metric_ids]

    preview_payload = build_preview_payload(payload, output)
    manifest_payload = build_manifest_payload(
        payload,
        preview_payload,
        output,
        preview_output,
        isolated_output,
        isolated_payload,
    )

    write_javascript_payload(output, "window.__footballScienceScoutingDatabase=", payload)
    write_javascript_payload(preview_output, "window.__footballScienceScoutingPreviewDatabase=", preview_payload)
    write_javascript_payload(manifest_output, "self.__footballScienceScoutingDatabaseManifest=", manifest_payload)
    write_javascript_payload(isolated_output, "self.__footballScienceNwslStatsbombDatabase=", isolated_payload)
    print(
        f"Wrote {output} with {len(records)} records and {len(metric_defs)} metrics; "
        f"isolated {len(isolated_payload['records'])} StatsBomb records in {isolated_output}."
    )


if __name__ == "__main__":
    main()
